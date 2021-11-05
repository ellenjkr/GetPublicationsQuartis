import pandas as pd
import requests
from requests.exceptions import HTTPError
import json
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from values import API_KEY

class Quartis():
	def __init__(self, file_name, api_key):
		super(Quartis, self).__init__()
		self.file_name = file_name
		self.api_key = api_key
		

	def search_percentil(self, issn):
		if issn != None:
			issn = issn.replace("-", "")
			issn = issn.replace(".0", "")
		percentil = None
		log = ''
		link_scopus = ''
		try:
			insttoken = os.environ.get('INSTTOKEN') # The insttoken is being hold as an environment variable
			headers = {'X-ELS-Insttoken': insttoken, 'X-ELS-APIKey': self.api_key}
			uri = "https://api.elsevier.com/content/serial/title?issn=" + issn + "&view=citescore"
			response = requests.get(uri, headers=headers)
			json_data = json.loads(response.text)

			link_scopus = json_data['serial-metadata-response']['entry'][0]['link'][0]['@href']
			try:
				percentil = json_data['serial-metadata-response']['entry'][0]['citeScoreYearInfoList']['citeScoreYearInfo'][1]['citeScoreInformationList'][0]['citeScoreInfo'][0]['citeScoreSubjectRank'][0]['percentile']
			except:
				log = 'sem valor de percentil'

		except HTTPError as http_err:
			log += 'HTTP error occurred: ' + str(http_err) + ' - Status_code: ' + str(response.status_code)
		except Exception as err:
			log += 'Other error occurred: ' + str(err) + ' - Status_code: ' + str(response.status_code)

		if percentil == '':
			percentil = None

		return percentil, link_scopus, str(log)


	def get_journal_str(self, main_journal):
		journal_array = main_journal.split(" ")

		journal_string = ""
		for i in range(len(journal_array)-1):
			journal_string += journal_array[i] + "+"
		journal_string += journal_array[-1]

		return journal_string

	def get_publication_page(self, journal_str, main_journal):
		r1 = requests.get(f"https://www.scimagojr.com/journalsearch.php?q={journal_str}")

		soup = BeautifulSoup(r1.text, 'lxml')
		soup = soup.find('div', class_='search_results')

		uri = None
		for link in soup.find_all('a'):
			journal_name = link.find_all(class_="jrnlname")
			for journal in journal_name:
				if journal.contents[0].lower() == main_journal.lower():
					uri = link.get('href')
		return uri

	def search_issn(self, uri):
		issn = "-"
		if uri != None:
			r2 = requests.get(f"https://www.scimagojr.com/{uri}")

			soup = BeautifulSoup(r2.text, 'lxml')
			soup = soup.find("div", class_="journaldescription colblock")

			table_items = soup.find_all("tr")

			
			for i in table_items:
				if str(i.contents[0]) == "<td>ISSN</td>":
					i.find_all("td")
					issn = i.contents[1].contents[0]

		return issn

	def get_issn(self, journal_name):
		journal_str = self.get_journal_str(journal_name)
		uri = self.get_publication_page(journal_str, journal_name)
		return self.search_issn(uri)
	

	def set_per_quarti(self, relatorio, value):
		per, link, log = self.search_percentil(issn=str(value))
		
		if per == None:
			print(value, " não foi possível")
			relatorio["Percentile"].append("-")
			relatorio["Quartile"].append("-")
		else:
			relatorio["Percentile"].append(per)
			if int(per) >= 75:
				quarti = "Q1"
			elif int(per) >= 50:
				quarti = "Q2"
			elif int(per) >= 25:
				quarti = "Q3"
			else:
				quarti = "Q4"

			relatorio["Quartile"].append(quarti)

		return relatorio


	def get_quartis(self):
		documents = pd.read_csv(self.file_name, sep=';', encoding='iso-8859-1')
		relatorio = {'Ano':[], 'Titulo':[], 'Nome de Publicação':[], 'ISSN':[], 'Percentile':[], 'Quartile':[]}

		for pos, value in enumerate(documents["issn"]):
			if value == "-":
				relatorio["Titulo"].append(documents["title"][pos])
				relatorio["Nome de Publicação"].append(documents["publication_name"][pos])

				issn = self.get_issn(documents["publication_name"][pos])
				relatorio["ISSN"].append(issn)

				if issn != "-":
					relatorio = self.set_per_quarti(relatorio, issn)
					relatorio["Ano"].append(documents["year"][pos])
				else:
					relatorio["Percentile"].append("-")
					relatorio["Quartile"].append("-")
					relatorio["Ano"].append(documents["year"][pos])
			else:
				relatorio["Titulo"].append(documents["title"][pos])
				relatorio["Nome de Publicação"].append(documents["publication_name"][pos])
				relatorio["ISSN"].append(value)

				relatorio = self.set_per_quarti(relatorio, value)

				relatorio["Ano"].append(documents["year"][pos])

		for pos, issn in enumerate(relatorio["ISSN"]):
			for pos2, issn2 in enumerate(relatorio["ISSN"]):
				if relatorio["Nome de Publicação"][pos] == relatorio["Nome de Publicação"][pos2]:
					if issn == '-' and issn2 != '-':
						relatorio["ISSN"][pos] = issn2
						relatorio["Percentile"][pos] = relatorio["Percentile"][pos2]
						relatorio["Quartile"][pos] = relatorio["Quartile"][pos2]

		return relatorio


class ExcelFile(Workbook):
	def __init__(self, relatorios):
		super(ExcelFile, self).__init__()
		self.relatorios = relatorios
		self.add_info()
		self.aplica_estilo()
		self.converte_valores()
		self.aplica_dimensoes()


	def add_info(self):
		for pos, autor in enumerate(self.relatorios["Autor"]):
			if pos == 0:
				ws = self.active # Primeiro sheet
				ws.title =  autor # Nome do sheet
			else:
				ws = self.create_sheet(autor)

			for row in dataframe_to_rows(self.relatorios["Relatorio"][pos], index=False, header=True): # Adiciona o dataframe ao sheet
			    ws.append(row)

	def aplica_estilo(self):
		for ws in self.worksheets:
			# Estilo dos nomes das colunas
			for cell in ws[1]:
				cell.font = Font(bold=True) # Negrito
				cell.alignment = Alignment(horizontal='center', vertical='center') # Centralizado

			# Centraliza todas as colunas menos a de título e de publicação
			for col in ws.columns:
				if col[0].column_letter not in 'BC':
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')

	def converte_valores(self):
		for ws in self.worksheets:
			for col in ws['D':'E']: # Converte colunas de ISSN e Percentile para número
				for cell in col:
					try:
						cell.value = int(cell.value)
					except:
						pass

	def aplica_dimensoes(self):
		for ws in self.worksheets:
			# Define tamanho das colunas
			ws.column_dimensions['A'].width = 15
			ws.column_dimensions['B'].width = 35
			ws.column_dimensions['C'].width = 35
			ws.column_dimensions['D'].width = 15
			ws.column_dimensions['E'].width = 15
			ws.column_dimensions['F'].width = 15

			ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas

	def salva_arquivo(self):
		self.save('quartis.xlsx') # Salva o arquivo
		

if __name__ == '__main__':
	relatorios = {'Autor':[], 'Relatorio':[]}

	for file in os.listdir("Autores"):
		quartis = Quartis(f"Autores/{file}", API_KEY')
		relatorio = quartis.get_quartis()
		relatorios['Relatorio'].append(pd.DataFrame(relatorio))
		relatorios['Autor'].append(file.split(".")[0])

	excel = ExcelFile(relatorios)
	excel.salva_arquivo()
