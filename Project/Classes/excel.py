from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelFile(Workbook):
	def __init__(self, relatorios):
		super(ExcelFile, self).__init__()
		self.relatorios = relatorios
		self.add_info()
		self.aplica_estilo()
		self.converte_valores()
		self.aplica_dimensoes()

	def add_info(self):
		for pos, autor in enumerate(self.relatorios["Autor"]):
			if pos == 0:
				ws = self.active  # Primeiro sheet
				ws.title = autor  # Nome do sheet
			else:
				ws = self.create_sheet(autor)

			for row in dataframe_to_rows(self.relatorios["Relatorio"][pos], index=False, header=True):  # Adiciona o dataframe ao sheet
			    ws.append(row)

	def aplica_estilo(self):
		for ws in self.worksheets:
			# Estilo dos nomes das colunas
			for cell in ws[1]:
				cell.font = Font(bold=True)  # Negrito
				cell.alignment = Alignment(horizontal='center', vertical='center')  # Centralizado

			# Centraliza todas as colunas menos a de título e de publicação
			for col in ws.columns:
				if col[0].column_letter not in 'BC':
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')

	def converte_valores(self):
		for ws in self.worksheets:
			for col in ws['D':'E']:  # Converte colunas de ISSN e Percentile para número
				for cell in col:
					try:
						cell.value = int(cell.value)
					except:
						pass

	def aplica_dimensoes(self):
		for ws in self.worksheets:
			# Define tamanho das colunas
			ws.column_dimensions['A'].width = 15
			ws.column_dimensions['B'].width = 35
			ws.column_dimensions['C'].width = 35
			ws.column_dimensions['D'].width = 15
			ws.column_dimensions['E'].width = 15
			ws.column_dimensions['F'].width = 15

			ws.row_dimensions[1].height = 25  # Altura da linha dos nomes das colunas

	def salva_arquivo(self):
		self.save('quartis.xlsx')  # Salva o arquivo